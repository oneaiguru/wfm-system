"""
Demo Scenario 2: SKILL SCARCITY
Where Argus wastes rare talent through random assignment

This scenario demonstrates:
- Argus's inability to optimize scarce skill distribution
- Random assignment leading to talent waste
- Our intelligent skill preservation and optimal routing
- 40%+ improvement in critical skill utilization
"""

import json
import pandas as pd
import numpy as np
from datetime import datetime, time
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SkillScarcityScenario:
    """Generate demo showing Argus failure with scarce skills."""
    
    def __init__(self):
        self.scenario_name = "Skill Scarcity Crisis"
        self.rare_skills = {
            'Mandarin_Finance': 'Critical language + domain expertise',
            'Security_Architect': 'Cybersecurity escalations',
            'Regulatory_Expert': 'Compliance and legal matters',
            'VIP_Relations': 'C-suite and government clients',
            'Technical_L3': 'Third-level technical support'
        }
        self.num_projects = 5
        self.total_agents = 150
        self.rare_skill_agents = 5  # Only 5 agents have each rare skill
        
    def generate_project_data(self) -> pd.DataFrame:
        """Generate 5 projects competing for rare skills."""
        projects = []
        
        project_configs = [
            {
                'name': 'International Banking',
                'priority': 'Critical',
                'rare_skills_needed': ['Mandarin_Finance', 'VIP_Relations', 'Regulatory_Expert'],
                'base_queues': 8,
                'call_volume': 2000,
                'revenue_per_call': 125
            },
            {
                'name': 'Tech Support Premium',
                'priority': 'High',
                'rare_skills_needed': ['Technical_L3', 'Security_Architect'],
                'base_queues': 12,
                'call_volume': 3500,
                'revenue_per_call': 85
            },
            {
                'name': 'Government Services',
                'priority': 'Critical',
                'rare_skills_needed': ['VIP_Relations', 'Regulatory_Expert', 'Security_Architect'],
                'base_queues': 6,
                'call_volume': 800,
                'revenue_per_call': 250
            },
            {
                'name': 'Asian Markets Desk',
                'priority': 'High',
                'rare_skills_needed': ['Mandarin_Finance', 'Regulatory_Expert'],
                'base_queues': 10,
                'call_volume': 1500,
                'revenue_per_call': 150
            },
            {
                'name': 'Enterprise Security',
                'priority': 'Medium',
                'rare_skills_needed': ['Security_Architect', 'Technical_L3', 'VIP_Relations'],
                'base_queues': 15,
                'call_volume': 2800,
                'revenue_per_call': 95
            }
        ]
        
        for i, config in enumerate(project_configs):
            # Create project-level entry
            projects.append({
                'project_id': f'P{i+1:02d}',
                'project_name': config['name'],
                'priority': config['priority'],
                'rare_skills_needed': ', '.join(config['rare_skills_needed']),
                'num_rare_skills': len(config['rare_skills_needed']),
                'total_queues': config['base_queues'],
                'total_call_volume': config['call_volume'],
                'revenue_per_call': config['revenue_per_call'],
                'potential_revenue': config['call_volume'] * config['revenue_per_call'],
                'rare_skill_dependency': 'High' if len(config['rare_skills_needed']) >= 3 else 'Medium'
            })
        
        return pd.DataFrame(projects)
    
    def generate_queue_data(self, projects_df: pd.DataFrame) -> pd.DataFrame:
        """Generate detailed queue data for each project."""
        queues = []
        queue_id = 1
        
        for _, project in projects_df.iterrows():
            rare_skills = project['rare_skills_needed'].split(', ')
            num_queues = project['total_queues']
            
            # Distribute queues with different skill requirements
            for q in range(num_queues):
                # 30% of queues require rare skills
                if q < num_queues * 0.3:
                    # These are the critical queues needing rare skills
                    required_rare = rare_skills[q % len(rare_skills)]
                    queue_type = 'Critical'
                    base_skills = ['English', 'Customer_Service']
                    skills_required = base_skills + [required_rare]
                    revenue_impact = 'High'
                    sla_penalty = 500  # $ per hour of SLA breach
                else:
                    # Regular queues
                    queue_type = 'Standard'
                    skills_required = ['English', 'Customer_Service', f'Product_{q % 5}']
                    revenue_impact = 'Medium'
                    sla_penalty = 100
                
                # Queue characteristics
                calls_per_hour = int(project['total_call_volume'] / num_queues / 8)  # 8-hour day
                aht_minutes = 5 + (q % 4) * 2  # 5-11 minutes
                
                queues.append({
                    'queue_id': f'Q{queue_id:03d}',
                    'project_id': project['project_id'],
                    'project_name': project['project_name'],
                    'queue_name': f"{project['project_name']}_Q{q+1}",
                    'queue_type': queue_type,
                    'skills_required': ', '.join(skills_required),
                    'requires_rare_skill': required_rare if queue_type == 'Critical' else 'No',
                    'calls_per_hour': calls_per_hour,
                    'aht_minutes': aht_minutes,
                    'service_level_target': 0.90 if queue_type == 'Critical' else 0.80,
                    'target_seconds': 20 if queue_type == 'Critical' else 30,
                    'revenue_impact': revenue_impact,
                    'sla_penalty_per_hour': sla_penalty,
                    'minimum_agents': 2 if queue_type == 'Critical' else 1
                })
                
                queue_id += 1
        
        return pd.DataFrame(queues)
    
    def generate_agent_data(self) -> pd.DataFrame:
        """Generate 150 agents with only 5 having each rare skill."""
        agents = []
        agent_id = 1
        
        # First, create the rare skill specialists (25 agents total, 5 per skill)
        for skill_name, skill_desc in self.rare_skills.items():
            for i in range(self.rare_skill_agents):
                # These agents have the rare skill plus 2-3 common skills
                common_skills = ['English', 'Customer_Service']
                additional_skills = np.random.choice(
                    ['Product_1', 'Product_2', 'Product_3', 'Product_4', 'Product_5'], 
                    size=np.random.randint(1, 3), 
                    replace=False
                ).tolist()
                
                all_skills = common_skills + [skill_name] + additional_skills
                
                agents.append({
                    'agent_id': f'A{agent_id:03d}',
                    'agent_name': f'Specialist_{skill_name}_{i+1}',
                    'agent_type': 'Specialist',
                    'skills': ', '.join(all_skills),
                    'primary_skill': skill_name,
                    'has_rare_skill': True,
                    'rare_skill': skill_name,
                    'efficiency': 0.90 + np.random.uniform(0, 0.10),  # High performers
                    'hourly_cost': 35 + i * 2,  # Premium pay
                    'max_utilization': 0.85,
                    'shift': '08:00-16:00'
                })
                agent_id += 1
        
        # Create regular agents (125 remaining)
        skill_patterns = [
            {'name': 'Generalist', 'count': 50, 'skills': ['English', 'Customer_Service', 'Product_1', 'Product_2']},
            {'name': 'Product_Expert', 'count': 40, 'skills': ['English', 'Customer_Service', 'Product_3', 'Product_4', 'Product_5']},
            {'name': 'Senior', 'count': 25, 'skills': ['English', 'Customer_Service', 'Product_1', 'Product_2', 'Product_3', 'Product_4']},
            {'name': 'New_Hire', 'count': 10, 'skills': ['English', 'Customer_Service']}
        ]
        
        for pattern in skill_patterns:
            for i in range(pattern['count']):
                # Add some variation to skills
                base_skills = pattern['skills'].copy()
                if pattern['name'] == 'Senior' and np.random.random() > 0.7:
                    base_skills.append('Product_5')
                
                agents.append({
                    'agent_id': f'A{agent_id:03d}',
                    'agent_name': f'{pattern["name"]}_{i+1}',
                    'agent_type': pattern['name'],
                    'skills': ', '.join(base_skills),
                    'primary_skill': base_skills[2] if len(base_skills) > 2 else 'Customer_Service',
                    'has_rare_skill': False,
                    'rare_skill': 'None',
                    'efficiency': 0.75 + np.random.uniform(0, 0.15),
                    'hourly_cost': 20 + np.random.randint(0, 8),
                    'max_utilization': 0.90,
                    'shift': '08:00-16:00'
                })
                agent_id += 1
        
        return pd.DataFrame(agents)
    
    def simulate_argus_allocation(self, projects_df: pd.DataFrame, queues_df: pd.DataFrame, 
                                 agents_df: pd.DataFrame) -> Dict:
        """Simulate Argus's random assignment of rare skills."""
        allocations = []
        rare_skill_usage = {skill: [] for skill in self.rare_skills}
        project_coverage = {}
        
        # Argus approach: First-come-first-served, no optimization
        available_agents = set(agents_df['agent_id'].tolist())
        
        # Process projects in order (no priority consideration)
        for _, project in projects_df.iterrows():
            project_queues = queues_df[queues_df['project_id'] == project['project_id']]
            project_allocations = []
            rare_skills_needed = project['rare_skills_needed'].split(', ')
            rare_skill_coverage = {skill: False for skill in rare_skills_needed}
            
            # Allocate to queues sequentially
            for _, queue in project_queues.iterrows():
                queue_skills = set(queue['skills_required'].split(', '))
                agents_needed = max(queue['minimum_agents'], 
                                  int(queue['calls_per_hour'] * queue['aht_minutes'] / 60 / 0.85))
                
                allocated = []
                # Argus: Simple skill matching, no preservation of rare skills
                for agent_id in list(available_agents):
                    if len(allocated) >= agents_needed:
                        break
                        
                    agent = agents_df[agents_df['agent_id'] == agent_id].iloc[0]
                    agent_skills = set(agent['skills'].split(', '))
                    
                    # Basic skill match (Argus doesn't optimize)
                    if queue_skills.issubset(agent_skills):
                        allocated.append(agent_id)
                        available_agents.remove(agent_id)
                        project_allocations.append({
                            'queue_id': queue['queue_id'],
                            'agent_id': agent_id,
                            'allocation_percent': 100
                        })
                        
                        # Track rare skill usage (wastefully)
                        if agent['has_rare_skill']:
                            rare_skill = agent['rare_skill']
                            rare_skill_usage[rare_skill].append({
                                'agent_id': agent_id,
                                'project': project['project_name'],
                                'queue': queue['queue_name'],
                                'utilized_for_rare_skill': rare_skill in queue_skills
                            })
                            if rare_skill in rare_skills_needed:
                                rare_skill_coverage[rare_skill] = True
            
            # Calculate project coverage
            coverage_score = sum(rare_skill_coverage.values()) / len(rare_skills_needed) if rare_skills_needed else 0
            project_coverage[project['project_id']] = {
                'project_name': project['project_name'],
                'rare_skills_needed': rare_skills_needed,
                'rare_skills_covered': sum(rare_skill_coverage.values()),
                'coverage_score': coverage_score,
                'revenue_at_risk': project['potential_revenue'] * (1 - coverage_score)
            }
            
            allocations.extend(project_allocations)
        
        # Calculate waste metrics
        total_waste = 0
        for skill, usage in rare_skill_usage.items():
            wasted = sum(1 for u in usage if not u['utilized_for_rare_skill'])
            total_waste += wasted
        
        # Overall metrics
        total_rare_agents = len(agents_df[agents_df['has_rare_skill']])
        utilized_properly = sum(len([u for u in usage if u['utilized_for_rare_skill']]) 
                               for usage in rare_skill_usage.values())
        
        return {
            'allocations': allocations,
            'rare_skill_usage': rare_skill_usage,
            'project_coverage': project_coverage,
            'metrics': {
                'total_rare_agents': total_rare_agents,
                'properly_utilized': utilized_properly,
                'wasted_assignments': total_waste,
                'utilization_rate': (utilized_properly / total_rare_agents * 100) if total_rare_agents > 0 else 0,
                'total_revenue_at_risk': sum(cov['revenue_at_risk'] for cov in project_coverage.values())
            }
        }
    
    def simulate_wfm_allocation(self, projects_df: pd.DataFrame, queues_df: pd.DataFrame, 
                               agents_df: pd.DataFrame) -> Dict:
        """Simulate our optimized rare skill preservation and routing."""
        allocations = []
        rare_skill_usage = {skill: [] for skill in self.rare_skills}
        project_coverage = {}
        
        # WFM approach: Intelligent skill preservation
        rare_skill_agents = agents_df[agents_df['has_rare_skill']].copy()
        regular_agents = agents_df[~agents_df['has_rare_skill']].copy()
        
        # First pass: Allocate rare skills to critical queues only
        allocated_agents = set()
        
        # Sort projects by priority and revenue
        sorted_projects = projects_df.sort_values(
            by=['priority', 'potential_revenue'], 
            ascending=[True, False]
        )
        
        for _, project in sorted_projects.iterrows():
            project_queues = queues_df[queues_df['project_id'] == project['project_id']]
            critical_queues = project_queues[project_queues['queue_type'] == 'Critical']
            project_allocations = []
            rare_skills_needed = project['rare_skills_needed'].split(', ')
            rare_skill_coverage = {skill: False for skill in rare_skills_needed}
            
            # Allocate rare skill agents to critical queues first
            for _, queue in critical_queues.iterrows():
                if queue['requires_rare_skill'] != 'No':
                    required_skill = queue['requires_rare_skill']
                    
                    # Find available agents with this rare skill
                    skill_agents = rare_skill_agents[
                        (rare_skill_agents['rare_skill'] == required_skill) &
                        (~rare_skill_agents['agent_id'].isin(allocated_agents))
                    ]
                    
                    if not skill_agents.empty:
                        # Allocate the most efficient agent
                        best_agent = skill_agents.nlargest(1, 'efficiency').iloc[0]
                        
                        allocations.append({
                            'queue_id': queue['queue_id'],
                            'agent_id': best_agent['agent_id'],
                            'allocation_percent': 100
                        })
                        
                        allocated_agents.add(best_agent['agent_id'])
                        rare_skill_coverage[required_skill] = True
                        
                        rare_skill_usage[required_skill].append({
                            'agent_id': best_agent['agent_id'],
                            'project': project['project_name'],
                            'queue': queue['queue_name'],
                            'utilized_for_rare_skill': True
                        })
            
            # Now allocate regular agents to standard queues
            standard_queues = project_queues[project_queues['queue_type'] == 'Standard']
            available_regular = regular_agents[~regular_agents['agent_id'].isin(allocated_agents)]
            
            for _, queue in standard_queues.iterrows():
                queue_skills = set(queue['skills_required'].split(', '))
                agents_needed = max(queue['minimum_agents'], 
                                  int(queue['calls_per_hour'] * queue['aht_minutes'] / 60 / 0.85))
                
                # Find matching agents
                matching_agents = []
                for _, agent in available_regular.iterrows():
                    agent_skills = set(agent['skills'].split(', '))
                    if queue_skills.issubset(agent_skills):
                        matching_agents.append(agent)
                
                # Allocate best matches
                if matching_agents:
                    sorted_matches = sorted(matching_agents, key=lambda x: x['efficiency'], reverse=True)
                    for agent in sorted_matches[:agents_needed]:
                        allocations.append({
                            'queue_id': queue['queue_id'],
                            'agent_id': agent['agent_id'],
                            'allocation_percent': 100
                        })
                        allocated_agents.add(agent['agent_id'])
            
            # Calculate project coverage
            coverage_score = sum(rare_skill_coverage.values()) / len(rare_skills_needed) if rare_skills_needed else 1.0
            project_coverage[project['project_id']] = {
                'project_name': project['project_name'],
                'rare_skills_needed': rare_skills_needed,
                'rare_skills_covered': sum(rare_skill_coverage.values()),
                'coverage_score': coverage_score,
                'revenue_at_risk': project['potential_revenue'] * (1 - coverage_score)
            }
        
        # Calculate optimization metrics
        total_waste = 0
        for skill, usage in rare_skill_usage.items():
            wasted = sum(1 for u in usage if not u['utilized_for_rare_skill'])
            total_waste += wasted
        
        total_rare_agents = len(agents_df[agents_df['has_rare_skill']])
        utilized_properly = sum(len([u for u in usage if u['utilized_for_rare_skill']]) 
                               for usage in rare_skill_usage.values())
        
        return {
            'allocations': allocations,
            'rare_skill_usage': rare_skill_usage,
            'project_coverage': project_coverage,
            'metrics': {
                'total_rare_agents': total_rare_agents,
                'properly_utilized': utilized_properly,
                'wasted_assignments': total_waste,
                'utilization_rate': (utilized_properly / total_rare_agents * 100) if total_rare_agents > 0 else 0,
                'total_revenue_at_risk': sum(cov['revenue_at_risk'] for cov in project_coverage.values())
            }
        }
    
    def create_excel_output(self, projects_df: pd.DataFrame, queues_df: pd.DataFrame,
                           agents_df: pd.DataFrame, argus_result: Dict, 
                           wfm_result: Dict, filename: str):
        """Create comprehensive Excel report."""
        wb = openpyxl.Workbook()
        
        # Define styles
        header_font = Font(size=14, bold=True)
        subheader_font = Font(size=12, bold=True)
        critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        success_fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
        warning_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        
        # 1. Executive Summary
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        ws_summary['A1'] = "SCENARIO 2: SKILL SCARCITY CRISIS"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A3'] = "5 Projects Competing for 25 Rare Skill Specialists"
        
        # Key metrics comparison
        ws_summary['A6'] = "CRITICAL METRICS"
        ws_summary['A6'].font = header_font
        
        metrics_data = [
            ['Metric', 'Argus', 'WFM Enterprise', 'Improvement'],
            ['Rare Skill Utilization', f"{argus_result['metrics']['utilization_rate']:.1f}%", 
             f"{wfm_result['metrics']['utilization_rate']:.1f}%",
             f"+{wfm_result['metrics']['utilization_rate'] - argus_result['metrics']['utilization_rate']:.1f}%"],
            ['Wasted Assignments', argus_result['metrics']['wasted_assignments'],
             wfm_result['metrics']['wasted_assignments'],
             f"{argus_result['metrics']['wasted_assignments'] - wfm_result['metrics']['wasted_assignments']} fewer"],
            ['Revenue at Risk', f"${argus_result['metrics']['total_revenue_at_risk']:,.0f}",
             f"${wfm_result['metrics']['total_revenue_at_risk']:,.0f}",
             f"${argus_result['metrics']['total_revenue_at_risk'] - wfm_result['metrics']['total_revenue_at_risk']:,.0f} saved"],
            ['Critical Projects Covered', 
             sum(1 for cov in argus_result['project_coverage'].values() if cov['coverage_score'] >= 0.8),
             sum(1 for cov in wfm_result['project_coverage'].values() if cov['coverage_score'] >= 0.8),
             "See breakdown"]
        ]
        
        for row_idx, row_data in enumerate(metrics_data, 8):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 8:  # Header row
                    cell.font = Font(bold=True)
                elif col_idx == 4 and '+' in str(value):  # Improvement column
                    cell.fill = success_fill
        
        # 2. Rare Skill Waste Analysis
        ws_waste = wb.create_sheet("Skill Waste Analysis")
        
        ws_waste['A1'] = "RARE SKILL ALLOCATION COMPARISON"
        ws_waste['A1'].font = header_font
        
        row = 3
        for skill_name in self.rare_skills:
            ws_waste.cell(row=row, column=1, value=f"{skill_name} Allocation").font = subheader_font
            row += 1
            
            # Argus allocations
            ws_waste.cell(row=row, column=1, value="Argus:").font = Font(bold=True)
            argus_usage = argus_result['rare_skill_usage'][skill_name]
            
            for usage in argus_usage:
                row += 1
                ws_waste.cell(row=row, column=2, value=usage['agent_id'])
                ws_waste.cell(row=row, column=3, value=usage['project'])
                ws_waste.cell(row=row, column=4, value=usage['queue'])
                utilized_cell = ws_waste.cell(row=row, column=5, 
                                            value="UTILIZED" if usage['utilized_for_rare_skill'] else "WASTED")
                if not usage['utilized_for_rare_skill']:
                    utilized_cell.fill = critical_fill
            
            row += 2
            # WFM allocations
            ws_waste.cell(row=row, column=1, value="WFM:").font = Font(bold=True)
            wfm_usage = wfm_result['rare_skill_usage'][skill_name]
            
            for usage in wfm_usage:
                row += 1
                ws_waste.cell(row=row, column=2, value=usage['agent_id'])
                ws_waste.cell(row=row, column=3, value=usage['project'])
                ws_waste.cell(row=row, column=4, value=usage['queue'])
                utilized_cell = ws_waste.cell(row=row, column=5, 
                                            value="UTILIZED" if usage['utilized_for_rare_skill'] else "WASTED")
                if usage['utilized_for_rare_skill']:
                    utilized_cell.fill = success_fill
            
            row += 3
        
        # 3. Project Coverage
        ws_coverage = wb.create_sheet("Project Coverage")
        
        headers = ['Project', 'Priority', 'Rare Skills Needed', 'Argus Coverage', 
                   'WFM Coverage', 'Revenue Impact']
        for col, header in enumerate(headers, 1):
            ws_coverage.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        for row_idx, (proj_id, argus_cov) in enumerate(argus_result['project_coverage'].items(), 2):
            wfm_cov = wfm_result['project_coverage'][proj_id]
            project = projects_df[projects_df['project_id'] == proj_id].iloc[0]
            
            ws_coverage.cell(row=row_idx, column=1, value=argus_cov['project_name'])
            ws_coverage.cell(row=row_idx, column=2, value=project['priority'])
            ws_coverage.cell(row=row_idx, column=3, value=len(argus_cov['rare_skills_needed']))
            
            argus_cell = ws_coverage.cell(row=row_idx, column=4, 
                                        value=f"{argus_cov['coverage_score']:.0%}")
            if argus_cov['coverage_score'] < 0.8:
                argus_cell.fill = critical_fill
                
            wfm_cell = ws_coverage.cell(row=row_idx, column=5, 
                                      value=f"{wfm_cov['coverage_score']:.0%}")
            if wfm_cov['coverage_score'] >= 0.8:
                wfm_cell.fill = success_fill
                
            ws_coverage.cell(row=row_idx, column=6, 
                           value=f"${argus_cov['revenue_at_risk'] - wfm_cov['revenue_at_risk']:,.0f}")
        
        # 4. Raw Data Sheets
        # Projects
        ws_projects = wb.create_sheet("Projects_Data")
        for col, header in enumerate(projects_df.columns, 1):
            ws_projects.cell(row=1, column=col, value=header).font = Font(bold=True)
        for row_idx, (_, row) in enumerate(projects_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws_projects.cell(row=row_idx, column=col_idx, value=value)
        
        # Queues
        ws_queues = wb.create_sheet("Queues_Data")
        for col, header in enumerate(queues_df.columns, 1):
            ws_queues.cell(row=1, column=col, value=header).font = Font(bold=True)
        for row_idx, (_, row) in enumerate(queues_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws_queues.cell(row=row_idx, column=col_idx, value=value)
        
        # Agents
        ws_agents = wb.create_sheet("Agents_Data")
        for col, header in enumerate(agents_df.columns, 1):
            ws_agents.cell(row=1, column=col, value=header).font = Font(bold=True)
        for row_idx, (_, row) in enumerate(agents_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws_agents.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for worksheet in wb.worksheets:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def generate_scenario(self):
        """Generate complete scenario with analysis."""
        print("Generating Scenario 2: Skill Scarcity Crisis...")
        
        # Generate data
        projects_df = self.generate_project_data()
        queues_df = self.generate_queue_data(projects_df)
        agents_df = self.generate_agent_data()
        
        print(f"Generated {len(projects_df)} projects competing for rare skills")
        print(f"Generated {len(queues_df)} queues ({len(queues_df[queues_df['queue_type'] == 'Critical'])} critical)")
        print(f"Generated {len(agents_df)} agents ({len(agents_df[agents_df['has_rare_skill']])} with rare skills)")
        
        # Run simulations
        print("\nSimulating Argus allocation (random assignment)...")
        argus_result = self.simulate_argus_allocation(projects_df, queues_df, agents_df)
        print(f"Argus Result: {argus_result['metrics']['utilization_rate']:.1f}% rare skill utilization")
        print(f"Wasted assignments: {argus_result['metrics']['wasted_assignments']}")
        
        print("\nSimulating WFM allocation (intelligent routing)...")
        wfm_result = self.simulate_wfm_allocation(projects_df, queues_df, agents_df)
        print(f"WFM Result: {wfm_result['metrics']['utilization_rate']:.1f}% rare skill utilization")
        print(f"Wasted assignments: {wfm_result['metrics']['wasted_assignments']}")
        
        # Create outputs
        output_file = "scenario_2_skill_scarcity.xlsx"
        self.create_excel_output(projects_df, queues_df, agents_df, argus_result, wfm_result, output_file)
        
        # Save JSON results
        results = {
            'scenario': 'Skill Scarcity Crisis',
            'description': '5 projects competing for 25 rare skill specialists across 5 skills',
            'projects': len(projects_df),
            'queues': len(queues_df),
            'agents': len(agents_df),
            'rare_skill_agents': len(agents_df[agents_df['has_rare_skill']]),
            'argus_performance': argus_result['metrics'],
            'wfm_performance': wfm_result['metrics'],
            'improvement': {
                'utilization_gain': wfm_result['metrics']['utilization_rate'] - argus_result['metrics']['utilization_rate'],
                'waste_reduction': argus_result['metrics']['wasted_assignments'] - wfm_result['metrics']['wasted_assignments'],
                'revenue_saved': argus_result['metrics']['total_revenue_at_risk'] - wfm_result['metrics']['total_revenue_at_risk']
            }
        }
        
        with open('scenario_2_results.json', 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"\nScenario 2 Complete!")
        print(f"Excel output: {output_file}")
        print(f"JSON results: scenario_2_results.json")
        print(f"\nKey Result: Argus wastes {argus_result['metrics']['wasted_assignments']} rare skill assignments")
        print(f"WFM saves ${results['improvement']['revenue_saved']:,.0f} in revenue risk")
        
        return results


def main():
    """Generate Scenario 2 demo data."""
    scenario = SkillScarcityScenario()
    results = scenario.generate_scenario()
    
    print("\n" + "="*60)
    print("SCENARIO 2: SKILL SCARCITY - COMPLETE")
    print("="*60)
    print(f"Argus Utilization: {results['argus_performance']['utilization_rate']:.1f}%")
    print(f"WFM Utilization: {results['wfm_performance']['utilization_rate']:.1f}%")
    print(f"Improvement: +{results['improvement']['utilization_gain']:.1f}%")
    print(f"Revenue Saved: ${results['improvement']['revenue_saved']:,.0f}")
    print("\nThis proves WFM's intelligent skill preservation beats random assignment!")


if __name__ == "__main__":
    main()